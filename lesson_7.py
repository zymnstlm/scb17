# -*- coding: utf-8 -*-
# @Author   :   YaMeng
# @File :   lesson_7.py
# @Software :   PyCharm
# @Time :   2020/11/20 14:13
# @company  :   湖南省零檬信息技术有限公司

# 自动化的步骤：
'''
1、excel准备好测试用例，并且通过代码读取到excel的测试用例   -- read_data()
2、发送接口请求，得到响应结果    --
3、执行结果  vs  预期结果
4、写入断言结果到excel
'''
import openpyxl
import requests

# 读取测试用例
def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]   #获取sheet
    max_row = sheet.max_row # 获取sheet里最大的行数
    # columns = sh.max_column  # 获取总列数
    list_1 = []  # 定义一个空列表，来接收所有的测试数据
    for i in range(2, max_row+1):  # 取头不取尾  左闭右开
        dict_1 = dict(
        id = sheet.cell(row=i, column=1).value,  # 取出id
        url = sheet.cell(row=i, column=5).value,  # 取出url
        data = sheet.cell(row=i, column=6).value,  # 取出data
        expect = sheet.cell(row=i, column=7).value)  # 取出expect
        # print(id,url,data,expect)
        list_1.append(dict_1)   # 把所有的测试数据，一一的追加到列表里
    return list_1
# 发送请求
def api_func(url, data):
    header = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    res1 = requests.post(url=url, json=data, headers=header)
    response = res1.json()
    return response
# 写入断言结果
def write_result(filename, sheetname, row, column, final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result
    wb.save(filename)
# 执行接口自动化测试
def execute_func(filename, sheetname):
    cases = read_data(filename, sheetname)
    for case in cases:
        id = case.get('id')   # 取出id
        url = case.get('url') # 取出url
        data = case.get('data') # 取出请求参数
        expect = case.get('expect') # 取出预期结果
        expect = eval(expect)
        expect_msg = expect.get('msg')  # 取出预期结果里面的msg信息
    # 从excel取出来的数据，都是str
        data = eval(data)   #  eval()运行被字符串包裹的python表达式 ==> 字符串转换为字典
        real_result = api_func(url=url, data=data)
        real_msg = real_result.get('msg')
        print('执行结果为：{}'.format(real_msg))
        print('预期结果为：{}'.format(expect_msg))
        if expect_msg == real_msg:
            print('这条测试用例通过！！')
            final_res = 'pass'
        else:
            print('这条测试用例不通过！！！')
            final_res = 'fail'
        print('*' * 30)
        write_result(filename, sheetname, id+1, 8, final_res)

execute_func('test_case_api.xlsx', 'register')
execute_func('test_case_api.xlsx', 'login')